#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
从 book.xlsx、person.xlsx、policy.xlsx 生成 SQLite 数据库 gujian_analytics.sqlite。
表名：books、persons、policies（各对应同名 Excel，policy 表来自 policy.xlsx）。
首行作为列名；若某文件不存在则写入示例 Excel 再导入。
"""

from __future__ import annotations

import json
import re
import shutil
import sqlite3
from pathlib import Path

try:
    from openpyxl import Workbook, load_workbook
except ImportError as e:
    raise SystemExit(
        "请先安装依赖：python -m pip install -r requirements.txt\n" + str(e)
    ) from e

DIR = Path(__file__).resolve().parent
BOOK_XLSX = DIR / "book.xlsx"
PERSON_XLSX = DIR / "person.xlsx"
POLICY_XLSX = DIR / "policy.xlsx"
DB_PATH = DIR / "gujian_analytics.sqlite"
# 人物肖像源目录（与 person.xlsx 同级下的 images/）；构建时同步到 public/images
PERSON_IMAGES_SRC = DIR / "images"
PUBLIC_PERSON_IMAGES = DIR.parent / "public" / "images"
_PERSON_IMAGE_EXTS = {".jpg", ".jpeg", ".png", ".webp", ".gif", ".bmp"}
_PERSON_IMAGE_COLS = frozenset({"图片", "图片参考", "头像", "照片", "image", "photo"})


def _slug_col(name: object, index: int) -> str:
    if name is None or str(name).strip() == "":
        return f"col_{index}"
    s = str(name).strip()
    s = re.sub(r"\s+", "_", s)
    return s or f"col_{index}"


def _ensure_sample_book() -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "books"
    headers = ["书名", "著者", "朝代", "类别", "卷册", "版本/馆藏", "备注"]
    ws.append(headers)
    ws.append(
        [
            "营造法式",
            "李诫",
            "北宋",
            "官修工程典籍",
            "三十四卷",
            "传世刻本多种",
            "古建筑设计与施工重要文献",
        ]
    )
    ws.append(
        [
            "园冶",
            "计成",
            "明",
            "园林",
            "三卷",
            "明崇祯刻本",
            "中国造园学专著",
        ]
    )
    ws.append(
        [
            "工程做法则例",
            "清工部",
            "清",
            "工程则例",
            "七十四卷",
            "清内府刊本",
            "清代官式建筑做法",
        ]
    )
    wb.save(BOOK_XLSX)


def _ensure_sample_person() -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "persons"
    headers = ["姓名", "字/号", "朝代", "籍贯", "职务/身份", "小传", "关联典籍", "图片"]
    ws.append(headers)
    ws.append(
        [
            "李诫",
            "字明仲",
            "北宋",
            "郑州管城",
            "将作监少监",
            "主持《营造法式》编纂",
            "营造法式",
            "",
        ]
    )
    ws.append(
        [
            "计成",
            "字无否",
            "明",
            "吴江",
            "造园家",
            "著《园冶》",
            "园冶",
            "",
        ]
    )
    ws.append(
        [
            "梁思成",
            "",
            "近现代",
            "广东新会",
            "建筑史学家",
            "中国古建筑学科奠基人之一",
            "营造法式（研究）",
            "",
        ]
    )
    wb.save(PERSON_XLSX)


def _ensure_sample_policy() -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "policies"
    headers = [
        "标题",
        "发文层次",
        "时期",
        "类别",
        "摘要",
        "关键词",
        "备注",
    ]
    ws.append(headers)
    ws.append(
        [
            "中华人民共和国文物保护法（节选要义）",
            "法律",
            "当代",
            "文物保护",
            "明确文物定义、保护原则、考古与修缮管理、法律责任等框架。",
            "文物;修缮;考古",
            "请以正式颁布文本为准，本表为结构示例。",
        ]
    )
    ws.append(
        [
            "历史文化名城名镇名村保护条例（摘要）",
            "行政法规",
            "当代",
            "城乡遗产",
            "规范历史文化名城、名镇、名村的申报、规划与监督管理。",
            "历史城区;风貌",
            "",
        ]
    )
    ws.append(
        [
            "文物建筑开放导则（要点）",
            "部门规范性文件",
            "当代",
            "活化利用",
            "引导文物建筑在确保安全前提下合理开放与展示。",
            "开放利用;展示",
            "",
        ]
    )
    wb.save(POLICY_XLSX)


def xlsx_to_sqlite(path: Path, table: str, conn: sqlite3.Connection) -> int:
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    rows = ws.iter_rows(values_only=True)
    raw_headers = next(rows, None)
    if not raw_headers:
        wb.close()
        return 0

    headers = [_slug_col(h, i) for i, h in enumerate(raw_headers)]
    col_sql = ", ".join(f'"{h}" TEXT' for h in headers)
    conn.execute(f"DROP TABLE IF EXISTS {table}")
    conn.execute(
        f"CREATE TABLE {table} (id INTEGER PRIMARY KEY AUTOINCREMENT, {col_sql})"
    )
    placeholders = ", ".join(["?"] * len(headers))
    quoted = ", ".join(f'"{h}"' for h in headers)
    insert_sql = f"INSERT INTO {table} ({quoted}) VALUES ({placeholders})"

    n = 0
    for row in rows:
        if row is None:
            continue
        vals = list(row[: len(headers)])
        while len(vals) < len(headers):
            vals.append(None)
        vals = [("" if v is None else str(v)) for v in vals]
        if all(v == "" for v in vals):
            continue
        conn.execute(insert_sql, vals)
        n += 1

    wb.close()
    return n


def export_books_json(conn: sqlite3.Connection) -> Path | None:
    """
    导出 books 表为前端静态 JSON，供数据分析圆盘「书目卡片」读取。
    输出：ancient_buildings/public/data/gujian_books.json
    """
    cur = conn.execute(
        "SELECT name FROM sqlite_master WHERE type='table' AND name='books'"
    )
    if not cur.fetchone():
        return None
    cur = conn.execute("PRAGMA table_info(books)")
    col_rows = cur.fetchall()
    cols = [r[1] for r in col_rows if r[1] != "id"]
    if not cols:
        return None
    quoted = ", ".join(f'"{c}"' for c in cols)
    rows = conn.execute(f"SELECT {quoted} FROM books").fetchall()
    books = [dict(zip(cols, row)) for row in rows]

    public_dir = DIR.parent / "public" / "data"
    public_dir.mkdir(parents=True, exist_ok=True)
    out = public_dir / "gujian_books.json"
    payload = {"version": 1, "columns": cols, "books": books}
    out.write_text(
        json.dumps(payload, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )
    return out


def export_policies_json(conn: sqlite3.Connection) -> Path | None:
    """
    导出 policies 表为前端静态 JSON，供「国家政策」等模块读取。
    输出：ancient_buildings/public/data/gujian_policies.json
    """
    cur = conn.execute(
        "SELECT name FROM sqlite_master WHERE type='table' AND name='policies'"
    )
    if not cur.fetchone():
        return None
    cur = conn.execute("PRAGMA table_info(policies)")
    col_rows = cur.fetchall()
    cols = [r[1] for r in col_rows if r[1] != "id"]
    if not cols:
        return None
    quoted = ", ".join(f'"{c}"' for c in cols)
    rows = conn.execute(f"SELECT {quoted} FROM policies").fetchall()
    policies = [dict(zip(cols, row)) for row in rows]

    public_dir = DIR.parent / "public" / "data"
    public_dir.mkdir(parents=True, exist_ok=True)
    out = public_dir / "gujian_policies.json"
    payload = {"version": 1, "columns": cols, "policies": policies}
    out.write_text(
        json.dumps(payload, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )
    return out


def sync_person_images_to_public() -> int:
    """将 数据分析/images 下的图片复制到 public/images，供前端以 /images/ 引用。"""
    if not PERSON_IMAGES_SRC.is_dir():
        return 0
    PUBLIC_PERSON_IMAGES.mkdir(parents=True, exist_ok=True)
    n = 0
    for f in PERSON_IMAGES_SRC.iterdir():
        if f.is_file() and f.suffix.lower() in _PERSON_IMAGE_EXTS:
            shutil.copy2(f, PUBLIC_PERSON_IMAGES / f.name)
            n += 1
    return n


def _portrait_basename_from_cell(val: object) -> str:
    s = ("" if val is None else str(val)).strip().replace("\\", "/")
    if not s or s.startswith("http://") or s.startswith("https://"):
        return ""
    return Path(s).name


def normalize_person_portraits(persons: list[dict], cols: list[str]) -> None:
    """
    将人物 JSON 中的肖像路径统一为 images/文件名（当 public/images 下存在该文件时）。
    若单元格为空，则尝试用「姓名」+ 常见后缀匹配目录中的文件。
    """
    if not PUBLIC_PERSON_IMAGES.is_dir():
        return
    available = {
        f.name
        for f in PUBLIC_PERSON_IMAGES.iterdir()
        if f.is_file() and f.suffix.lower() in _PERSON_IMAGE_EXTS
    }
    if not available:
        return

    img_keys = [k for k in cols if k in _PERSON_IMAGE_COLS]
    if not img_keys:
        return
    primary_key = img_keys[0]

    def assign(p: dict, basename: str) -> None:
        if not basename or basename not in available:
            return
        path = f"images/{basename}"
        for k in img_keys:
            if k in p:
                p[k] = path
                return
        p[primary_key] = path

    for p in persons:
        matched = False
        for k in img_keys:
            base = _portrait_basename_from_cell(p.get(k, ""))
            if base and base in available:
                p[k] = f"images/{base}"
                matched = True
                break
        if matched:
            continue
        name = (p.get("姓名") or "").strip()
        short = name.split("（")[0].strip() if name else ""
        if not short:
            continue
        for ext in (".jpg", ".jpeg", ".png", ".webp", ".gif"):
            fn = short + ext
            if fn in available:
                assign(p, fn)
                break


def export_persons_json(conn: sqlite3.Connection) -> Path | None:
    """
    导出 persons 表为前端静态 JSON，供「栋宇良工」人物时间线读取。
    输出：ancient_buildings/public/data/gujian_persons.json
    """
    cur = conn.execute(
        "SELECT name FROM sqlite_master WHERE type='table' AND name='persons'"
    )
    if not cur.fetchone():
        return None
    cur = conn.execute("PRAGMA table_info(persons)")
    col_rows = cur.fetchall()
    cols = [r[1] for r in col_rows if r[1] != "id"]
    if not cols:
        return None
    quoted = ", ".join(f'"{c}"' for c in cols)
    rows = conn.execute(f"SELECT {quoted} FROM persons").fetchall()
    persons = [dict(zip(cols, row)) for row in rows]

    sync_person_images_to_public()
    normalize_person_portraits(persons, cols)

    public_dir = DIR.parent / "public" / "data"
    public_dir.mkdir(parents=True, exist_ok=True)
    out = public_dir / "gujian_persons.json"
    payload = {"version": 1, "columns": cols, "persons": persons}
    out.write_text(
        json.dumps(payload, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )
    return out


def main() -> None:
    if not BOOK_XLSX.is_file():
        print("未找到 book.xlsx，已生成示例文件：", BOOK_XLSX)
        _ensure_sample_book()
    if not PERSON_XLSX.is_file():
        print("未找到 person.xlsx，已生成示例文件：", PERSON_XLSX)
        _ensure_sample_person()
    if not POLICY_XLSX.is_file():
        print("未找到 policy.xlsx，已生成示例文件：", POLICY_XLSX)
        _ensure_sample_policy()

    conn = sqlite3.connect(DB_PATH)
    try:
        n_book = xlsx_to_sqlite(BOOK_XLSX, "books", conn)
        n_person = xlsx_to_sqlite(PERSON_XLSX, "persons", conn)
        n_policy = xlsx_to_sqlite(POLICY_XLSX, "policies", conn)
        conn.commit()
        print(f"已写入 {DB_PATH}")
        print(f"  books    行数（不含表头）: {n_book}")
        print(f"  persons  行数（不含表头）: {n_person}")
        print(f"  policies 行数（不含表头）: {n_policy}")
        jp = export_books_json(conn)
        if jp:
            print(f"  已导出前端书目 JSON：{jp}")
        jpp = export_persons_json(conn)
        if jpp:
            print(f"  已导出前端人物 JSON：{jpp}")
            if PERSON_IMAGES_SRC.is_dir():
                print("  人物图片目录：数据分析/images → 已同步至 public/images")
        jpol = export_policies_json(conn)
        if jpol:
            print(f"  已导出前端政策 JSON：{jpol}")
    finally:
        conn.close()


if __name__ == "__main__":
    main()
